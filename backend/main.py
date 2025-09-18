from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
import pandas as pd
import os
from datetime import datetime
from typing import List, Optional
import io

app = FastAPI(title="Health Survey API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data models
class Employee(BaseModel):
    EmployeeNumber: str
    EmployeeName: str

class SurveySubmission(BaseModel):
    employeeNumber: str
    employeeName: str
    gender: str
    age: int
    waistCircumference: float
    heightFeet: int
    heightInches: int
    weight: float
    bmi: float
    bmiCategory: str

# File paths
EMPLOYEES_CSV = "data/employees.csv"
SURVEY_DATA_CSV = "data/survey_responses.csv"

# Ensure data directory exists
os.makedirs("data", exist_ok=True)

def load_employees() -> List[Employee]:
    """Load employee data from CSV with proper string handling for EmployeeNumber"""
    try:
        if not os.path.exists(EMPLOYEES_CSV):
            raise HTTPException(status_code=404, detail="Employee data file not found")
        
        # Read CSV with EmployeeNumber as string to preserve leading zeros
        df = pd.read_csv(EMPLOYEES_CSV, dtype={'EmployeeNumber': str})
        
        employees = []
        for _, row in df.iterrows():
            # Ensure EmployeeNumber is properly formatted with leading zeros
            emp_num = str(row['EmployeeNumber']).zfill(8)  # Pad to 8 digits
            employees.append(Employee(
                EmployeeNumber=emp_num,
                EmployeeName=str(row['EmployeeName'])
            ))
        return employees
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading employee data: {str(e)}")

def initialize_survey_csv():
    """Initialize survey responses CSV with proper column names matching input fields"""
    if not os.path.exists(SURVEY_DATA_CSV):
        # Column names that match exactly with the frontend input fields
        headers = [
            'SubmissionDate',
            'EmployeeNumber', 
            'EmployeeName',
            'Gender',
            'Age',
            'Waist Circumference (inches)',
            'Height - Feet',
            'Height - Inches', 
            'Weight (lb)',
            'BMI',
            'BMI Category'
        ]
        # Create empty DataFrame with headers and save it
        empty_df = pd.DataFrame(columns=headers)
        empty_df.to_csv(SURVEY_DATA_CSV, index=False)
        print(f"Initialized survey CSV with headers: {headers}")

@app.on_event("startup")
async def startup_event():
    """Initialize CSV files on startup"""
    initialize_survey_csv()
    
    # Create sample employee data if it doesn't exist
    if not os.path.exists(EMPLOYEES_CSV):
        sample_data = '''EmployeeNumber,EmployeeName
"00071215","Pyae Phyo Latt"
"00070782","Ni Ni Aung"
"00071156","Ayar Lwin"
"00070098","Aye Aye Tun"
"00071182","Myo Min Min Wai"
"00070039","Kyaw Zayar Myint"
"00070961","Min Soe Moe Naung"
"00070671","Tin Maung Htwe"
"00070459","San San Aung"
"00070081","Nilar"
"00070291","Ohnmar"
"00070783","Ko Ko Zin"
"00070314","Zaw Zaw Oo @ Win Aung"
"00070036","Theingi Aung"
"00070618","Min Hein Kyaw"
"00071216","Kyaw Soe Moe"
"00071157","Nwet Nwet San"
"00070598","Tin Tun Aung"
"00070292","Than Than Nwet"
"00070129","Thandar Win"
"00070725","Mo Mo Phone Maw @ Moh Moh Naing"
"00070126","Su Sandy Linn"
"00071168","Aung Ko Ko"
"00070798","Htin Lin Oo"
"00071214","Zin Ko Myint"
"00071239","Aung Thu Lwin"
"00070785","Ko Ko Htet"
"00070636","Nyein Nyein Hlaing"
"00070792","Zaw Win"
"00071577","Phyo Wai Maung"
"00070944","Than Htike Oo"
"00070128","Khine Nyo Tun"
"00071201","Kyaw Soe Win"
"00070730","Yan Naing"
"00071131","Wai Mar"
"00071709","Htoo Yadanar Kyaw"
"00071002","Soe San"
"00070962","Zin Thu Ma Ma Aung"
"00071909","Aung Kyaw Soe"'''
        
        with open(EMPLOYEES_CSV, 'w') as f:
            f.write(sample_data)

@app.get("/")
async def root():
    return {"message": "Health Survey API is running"}

@app.get("/employees", response_model=List[Employee])
async def get_employees():
    """Get list of all employees"""
    return load_employees()

@app.post("/submit-survey")
async def submit_survey(survey: SurveySubmission):
    """Submit a new survey response with proper column names"""
    try:
        # Validate employee exists
        employees = load_employees()
        employee_numbers = [emp.EmployeeNumber for emp in employees]
        
        # Ensure submitted employee number has proper format
        formatted_emp_num = str(survey.employeeNumber).zfill(8)
        
        if formatted_emp_num not in employee_numbers:
            raise HTTPException(status_code=400, detail="Invalid employee number")
        
        # Prepare data for CSV with proper column names matching the headers
        survey_data = {
            'SubmissionDate': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'EmployeeNumber': formatted_emp_num,
            'EmployeeName': survey.employeeName,
            'Gender': survey.gender,
            'Age': survey.age,
            'Waist Circumference (inches)': survey.waistCircumference,
            'Height - Feet': survey.heightFeet,
            'Height - Inches': survey.heightInches,
            'Weight (lb)': survey.weight,
            'BMI': round(survey.bmi, 1),
            'BMI Category': survey.bmiCategory
        }
        
        # Check if file exists and has data
        file_exists = os.path.exists(SURVEY_DATA_CSV)
        
        if file_exists:
            # Read existing data
            existing_df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
            # Create new row DataFrame
            new_row_df = pd.DataFrame([survey_data])
            # Concatenate existing data with new row
            updated_df = pd.concat([existing_df, new_row_df], ignore_index=True)
        else:
            # Create new DataFrame with just this row
            updated_df = pd.DataFrame([survey_data])
        
        # Write the complete DataFrame back to CSV (this preserves headers)
        updated_df.to_csv(SURVEY_DATA_CSV, index=False)
        
        return {"message": "Survey submitted successfully", "data": survey_data}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error submitting survey: {str(e)}")

@app.get("/survey-responses")
async def get_survey_responses():
    """Get all survey responses"""
    try:
        if not os.path.exists(SURVEY_DATA_CSV):
            return []
        
        # Read with proper string handling for EmployeeNumber
        df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
        return df.to_dict('records')
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving survey responses: {str(e)}")

@app.get("/download-survey-responses")
async def download_survey_responses():
    """Download all survey responses as a CSV file with proper headers"""
    try:
        if not os.path.exists(SURVEY_DATA_CSV):
            raise HTTPException(status_code=404, detail="No survey responses found")
        
        # Read the survey data with proper column handling
        df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
        
        if df.empty:
            raise HTTPException(status_code=404, detail="No survey responses available for download")
        
        # Ensure column names are exactly what we want
        expected_columns = [
            'SubmissionDate',
            'EmployeeNumber', 
            'EmployeeName',
            'Gender',
            'Age',
            'Waist Circumference (inches)',
            'Height - Feet',
            'Height - Inches', 
            'Weight (lb)',
            'BMI',
            'BMI Category'
        ]
        
        # Reorder columns to match expected order
        df = df.reindex(columns=expected_columns)
        
        # Create a CSV string with headers
        output = io.StringIO()
        df.to_csv(output, index=False, quoting=1)  # quoting=1 preserves leading zeros
        csv_content = output.getvalue()
        
        # Debug: Print first few lines to console
        print("CSV Content (first 200 chars):")
        print(csv_content[:200])
        
        # Create filename with current timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"survey_responses_{timestamp}.csv"
        
        # Return as streaming response
        response = StreamingResponse(
            io.BytesIO(csv_content.encode('utf-8')),
            media_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
        return response
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Survey responses file not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading survey responses: {str(e)}")

@app.get("/download-survey-responses-excel")
async def download_survey_responses_excel():
    """Download all survey responses as an Excel file with proper headers"""
    try:
        if not os.path.exists(SURVEY_DATA_CSV):
            raise HTTPException(status_code=404, detail="No survey responses found")
        
        # Read the survey data
        df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
        
        if df.empty:
            raise HTTPException(status_code=404, detail="No survey responses available for download")
        
        # Ensure column names are properly formatted
        expected_columns = [
            'SubmissionDate',
            'EmployeeNumber', 
            'EmployeeName',
            'Gender',
            'Age',
            'Waist Circumference (inches)',
            'Height - Feet',
            'Height - Inches', 
            'Weight (lb)',
            'BMI',
            'BMI Category'
        ]
        
        # Reorder columns to match expected order
        df = df.reindex(columns=expected_columns)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the DataFrame to Excel with headers
            df.to_excel(writer, sheet_name='Survey Responses', index=False, startrow=0)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Survey Responses']
            
            # Format the header row
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to header row
            for cell in worksheet[1]:  # First row (header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 12), 50)  # Min width 12, max 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
            
            # Add borders to all data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
        
        output.seek(0)
        
        # Create filename with current timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"survey_responses_{timestamp}.xlsx"
        
        # Return as streaming response
        response = StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
        return response
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Survey responses file not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading survey responses: {str(e)}")


@app.get("/debug-csv")
async def debug_csv():
    """Debug endpoint to check CSV structure"""
    try:
        if not os.path.exists(SURVEY_DATA_CSV):
            return {"message": "CSV file doesn't exist"}
        
        df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
        
        return {
            "file_exists": True,
            "columns": list(df.columns),
            "total_rows": len(df),
            "sample_data": df.head(2).to_dict('records') if not df.empty else [],
            "file_size_bytes": os.path.getsize(SURVEY_DATA_CSV)
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/survey-stats")
async def get_survey_statistics():
    """Get basic statistics about survey responses"""
    try:
        if not os.path.exists(SURVEY_DATA_CSV):
            return {
                "total_responses": 0,
                "message": "No survey data available"
            }
        
        df = pd.read_csv(SURVEY_DATA_CSV, dtype={'EmployeeNumber': str})
        
        if df.empty:
            return {
                "total_responses": 0,
                "message": "No survey responses available"
            }
        
        # Calculate statistics using the proper column names
        total_responses = len(df)
        gender_distribution = df['Gender'].value_counts().to_dict() if 'Gender' in df.columns else {}
        avg_age = round(df['Age'].mean(), 1) if 'Age' in df.columns and not df['Age'].isna().all() else 0
        avg_bmi = round(df['BMI'].mean(), 1) if 'BMI' in df.columns and not df['BMI'].isna().all() else 0
        bmi_categories = df['BMI Category'].value_counts().to_dict() if 'BMI Category' in df.columns else {}
        
        # Calculate additional statistics
        avg_waist = round(df['Waist Circumference (inches)'].mean(), 1) if 'Waist Circumference (inches)' in df.columns and not df['Waist Circumference (inches)'].isna().all() else 0
        avg_weight = round(df['Weight (lb)'].mean(), 1) if 'Weight (lb)' in df.columns and not df['Weight (lb)'].isna().all() else 0
        
        # Calculate average height in feet and inches
        if 'Height - Feet' in df.columns and 'Height - Inches' in df.columns:
            total_height_inches = df['Height - Feet'] * 12 + df['Height - Inches']
            avg_height_total_inches = round(total_height_inches.mean(), 1) if not total_height_inches.isna().all() else 0
            avg_height_feet = int(avg_height_total_inches // 12) if avg_height_total_inches > 0 else 0
            avg_height_inches = round(avg_height_total_inches % 12, 1) if avg_height_total_inches > 0 else 0
        else:
            avg_height_feet = 0
            avg_height_inches = 0
        
        # Get date range
        if 'SubmissionDate' in df.columns:
            df['SubmissionDate'] = pd.to_datetime(df['SubmissionDate'], errors='coerce')
            first_response = df['SubmissionDate'].min().strftime('%Y-%m-%d %H:%M:%S') if not df['SubmissionDate'].isna().all() else "N/A"
            last_response = df['SubmissionDate'].max().strftime('%Y-%m-%d %H:%M:%S') if not df['SubmissionDate'].isna().all() else "N/A"
        else:
            first_response = "N/A"
            last_response = "N/A"
        
        return {
            "total_responses": total_responses,
            "gender_distribution": gender_distribution,
            "averages": {
                "age": avg_age,
                "bmi": avg_bmi,
                "waist_circumference_inches": avg_waist,
                "weight_lb": avg_weight,
                "height_feet": avg_height_feet,
                "height_inches": avg_height_inches
            },
            "bmi_categories": bmi_categories,
            "date_range": {
                "first_response": first_response,
                "last_response": last_response
            },
            "file_info": {
                "size_kb": round(os.path.getsize(SURVEY_DATA_CSV) / 1024, 2) if os.path.exists(SURVEY_DATA_CSV) else 0,
                "location": SURVEY_DATA_CSV
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving survey statistics: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)